import os
import sys
import logging
import asyncio
import time
import asyncpraw.exceptions
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
import asyncpraw
from rate_limiter import APIRateLimiter
from openpyxl.worksheet.worksheet import Worksheet

load_dotenv(dotenv_path=".env")

logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

rate_limit = APIRateLimiter(logger, max_retries=3, initial_delay=1.0)


class RedditAPIClient:
    """
    A client for interacting with Reddit's API and processing submission data.

    Attributes:
        reddit (asyncpraw.Reddit): Reddit API client instance
        xlsxclient (ExcelHandler): Excel file handler for data I/O
    """

    def __init__(self, read_file_path: str, write_file_path: str) -> None:
        """
        Initialize the Reddit API client.

        Args:
            read_file_path: Path to input Excel file
            write_file_path: Path to output Excel file
        """
        self.reddit = asyncpraw.Reddit(
            client_id=os.getenv("CLIENT_ID"),
            client_secret=os.getenv("CLIENT_SECRET"),
            user_agent=os.getenv("USER_AGENT"),
        )
        self.xlsxclient = ExcelHandler(read_file_path, write_file_path)

    async def __aenter__(self) -> 'RedditAPIClient':
        return self

    async def __aexit__(self, exc_type: None | type, exc_val: None | BaseException, exc_tb: None | object) -> None:
        await self.reddit.close()

    @rate_limit.rate_limit
    async def get_submission_comments(self, submission_url: str, traffic: str) -> None:
        """
        Fetch and process comments for a given submission URL.
        If the submission is locked or archived, it is skipped.
        Args:
            submission_url: Reddit submission URL to process
            traffic: Traffic category of the submission
        """
        try:
            submission = await self.reddit.submission(url=submission_url)
        except asyncpraw.exceptions.InvalidURL:
            logger.warning(
                f"Invalid submission URL: {submission_url}. Skipping...")
            return
        if submission.locked or submission.archived:
            return
        comments_count = len(submission.comments)
        await self._write_comments_to_excel(submission_url, comments_count, traffic)
        logger.info(
            f"Successfully processed submission: {submission_url} with {comments_count} comments")

    async def _write_comments_to_excel(self, submission_url: str, comments_count: int, traffic: str) -> None:
        if comments_count == 0:
            await self.xlsxclient.write_data_to_sheet([submission_url, 0, traffic], "No comments")
        elif 1 <= comments_count <= 3:
            await self.xlsxclient.write_data_to_sheet([submission_url, comments_count, traffic], "3 or less comments")

    async def run(self, skip_header: bool = True) -> None:
        data = await self.xlsxclient.read_data()
        if data is None:
            logger.warning("No data to process in file!")
            return

        tasks = []
        for sheet_name, sheet_data in data.items():
            for index, row in enumerate(sheet_data):
                if skip_header and index == 0:
                    continue
                url, traffic = row
                tasks.append(self.get_submission_comments(url, traffic))
                logger.info(f"Queued: row {index} in sheet {sheet_name}")

        await asyncio.gather(*tasks)
        await self.xlsxclient.sort_output_by_traffic()
        logger.info("Processing complete!")


class ExcelHandler:
    """
    Handles Excel file operations for reading and writing Reddit data.

    Attributes:
        read_file_path (str): Path to input Excel file
        write_file_path (str): Path to output Excel file
        read_workbook: Loaded workbook for reading
        write_workbook: Workbook for writing results
    """

    def __init__(self, read_file_path: str, write_file_path: str) -> None:
        """
        Initialize the Excel handler.

        Args:
            read_file_path: Path to input Excel file
            write_file_path: Path to output Excel file
        """
        self.read_file_path = read_file_path
        self.write_file_path = write_file_path
        self.read_workbook: None | Workbook = None
        self.write_workbook: Workbook = Workbook()

    async def read_data(self) -> None | dict[str, list[list[str | int]]]:
        """
        Read data from the read_file_path Excel file.

        Returns:
            A dictionary where keys are sheet names and values are lists of rows, 
            with each row being a list of cell values. Returns None if an error occurs.
            
        """
        try:
            self.read_workbook = load_workbook(self.read_file_path)
            return await self._extract_data_from_workbook()
        except FileNotFoundError:
            logger.error(f"Error: File '{self.read_file_path}' not found.")
        except InvalidFileException:
            logger.error(
                "Error: Invalid file format. Please provide an Excel file.")
        except Exception as e:
            logger.error(f"Error reading file: {e}")

    async def _extract_data_from_workbook(self) -> dict[str, list[list[str | int]]]:
        """
        Extract data from the loaded workbook from each sheet.

        Returns:
            A dictionary where keys are sheet names and values are lists of rows, 
            with each row being a list of cell values.
        """
        data: dict[str, list[list[str | int]]] = {}
        for sheet_name in self.read_workbook.sheetnames:
            sheet = self.read_workbook[sheet_name]
            data[sheet_name] = [
                [cell.value for cell in row] for row in sheet.iter_rows()
            ]
        return data

    async def write_data_to_sheet(self, row_data: list[str | int], sheet_name: str) -> None:
        """
        Write a row of data to a specified sheet in the output Excel file.

        This method appends a row of data to the specified sheet in the write_workbook.
        If the sheet does not exist, it creates a new sheet with the given name.

        Args:
            row_data: The row data to write.
            sheet_name: The name of the sheet to write to.
        """
        sheet = await self._get_or_create_sheet(sheet_name)
        sheet.append(row_data)
        await self._save_workbook()

    async def _save_workbook(self) -> None:
        """
        Save the output Excel workbook.

        This method saves the write_workbook to the specified write_file_path.
        """
        self.write_workbook.save(self.write_file_path)

    async def _get_or_create_sheet(self, sheet_name: str) -> Worksheet:
        """
        Get an existing sheet or create a new one if it doesn't exist.

        This method checks if a sheet with the given name exists in the write_workbook.
        If it exists, it returns the sheet. Otherwise, it creates a new sheet with the given name.

        Args:
            sheet_name: The name of the sheet to get or create.

        Returns:
            The worksheet object.
        """
        if sheet_name in self.write_workbook.sheetnames:
            return self.write_workbook[sheet_name]
        else:
            sheet = self.write_workbook.create_sheet(sheet_name)
            sheet.append(["URL", "Number of comments", "Traffic"])
            return sheet

    async def sort_output_by_traffic(self) -> None:
        """
        Sort the data in each sheet of the output Excel file by traffic.

        This method iterates through each sheet in the write_workbook, sorts the data by the traffic column,
        replaces the existing data with the sorted data, and then saves the workbook.
        """
        for sheet_name in self.write_workbook.sheetnames:
            sheet = self.write_workbook[sheet_name]
            sorted_data = await self._sort_sheet_data_by_traffic(sheet)
            await self._replace_sheet_data(sheet, sorted_data)
        await self._save_workbook()

    async def _sort_sheet_data_by_traffic(self, sheet: Worksheet) -> list[tuple[str | int, ...]]:
        """
        Sort the data in a sheet by the traffic column.

        This method sorts the rows of data in the given sheet by the traffic column (assumed to be the third column).
        The header row is excluded from the sorting.

        Args:
            sheet: The worksheet to sort.

        Returns:
            A list of tuples representing the sorted rows of data.
        """
        return sorted((row for row in sheet.iter_rows(min_row=2, values_only=True)), key=lambda x: x[2], reverse=True)

    async def _replace_sheet_data(self, sheet: Worksheet, sorted_data: list[tuple[str | int, ...]]) -> None:
        """
        Replace the data in a sheet with sorted data.

        This method deletes all rows in the given sheet starting from the second row and appends the sorted data.

        Args:
            sheet: The worksheet to replace data in.
            sorted_data: The sorted data to append to the sheet.
        """
        sheet.delete_rows(2, sheet.max_row)
        for row in sorted_data:
            sheet.append(row)


async def main() -> None:
    start_time = time.time()
    try:
        async with RedditAPIClient(read_file_path=input_read_file_path,
                                   write_file_path=output_file_path) as reddit_script:
            await reddit_script.run()
    finally:
        end_time = time.time()
        execution_time = end_time - start_time
        logger.info(f"Total execution time: {execution_time:.2f} seconds")

if __name__ == "__main__":
    input_read_file_path = sys.argv[1]
    output_file_path = sys.argv[2]

    asyncio.run(main())
